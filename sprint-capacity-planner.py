# library modules
import datetime
import time
import openpyxl
import sys


# local include files
import credentials as cr
from scp_classes import Default, BankHoliday, Developer, Vacation, Sprint

# global variables
bank_holidays = []
vacations = []
developers = []
sprints = []
defaults = []

start_time = time.time()


def load_dictionary(full_file_name, sheet_name, load_max_col, load_max_row=50):
    xlsfile = full_file_name['path'] + '/' + full_file_name['filename']
    all_rows = []
    if load_max_col < 1:
        all_rows[0] = 'load_max_col must be greater than 0'
        return all_rows
    if load_max_row < 1:
        all_rows[0] = 'load_max_row must be greater than 0'
        return all_rows
    try:
        wb = openpyxl.load_workbook(xlsfile, read_only=True, data_only=True)
    except:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    keys = []
    try:
        sheet = wb[sheet_name]
    except:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    for firstrow in sheet.iter_rows(min_row=1, max_row=1, min_col=0, max_col=load_max_col, values_only=True):
        for key in firstrow:
            keys.append(key)
    for row in sheet.iter_rows(min_row=2, max_row=load_max_row, min_col=0, max_col=load_max_col, values_only=True):
        onerow = {}
        for i in range(0, load_max_col):
            onerow[firstrow[i]] = row[i]
        all_rows.append(onerow)
    return all_rows


def load_to_objects(full_file_name, sheet_name, load_max_col, load_max_row=50):
    xlsfile = full_file_name['path'] + '/' + full_file_name['filename']
    all_rows = []
    if load_max_col < 1:
        all_rows[0] = 'load_max_col must be greater than 0'
        return all_rows
    if load_max_row < 1:
        all_rows[0] = 'load_max_row must be greater than 0'
        return all_rows
    try:
        wb = openpyxl.load_workbook(xlsfile, read_only=True, data_only=True)
    except:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    keys = []
    try:
        sheet = wb[sheet_name]
    except:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    for firstrow in sheet.iter_rows(min_row=1, max_row=1, min_col=0, max_col=load_max_col, values_only=True):
        for key in firstrow:
            keys.append(key)
    for row in sheet.iter_rows(min_row=2, max_row=load_max_row, min_col=0, max_col=load_max_col, values_only=True):
        onerow = {}
        for i in range(0, load_max_col):
            onerow[firstrow[i]] = row[i]
        all_rows.append(onerow)
    return all_rows


def load_all():
    global bank_holidays
    global vacations
    global developers
    global sprints
    global defaults
    bank_holidays = load_dictionary(cr.testexcelfile, 'Bank holidays', 3, 100)
    vacations = load_dictionary(cr.testexcelfile, 'Vacations', 16, 200)
    developers = load_dictionary(cr.testexcelfile, 'Developers', 5)
    sprints = load_dictionary(cr.testexcelfile, 'Sprints', 7, 100)
    defaults = load_dictionary(cr.testexcelfile, 'Defaults', 2)


if __name__ == '__main__':
    now = datetime.datetime.now()
    print("Started at:", now.strftime("%Y-%m-%d %H:%M:%S"))
    print("#####----------------------------------#####")
    developer = Developer('Bela', 'HU', 1, None, None)
    print(developer)
    load_all()
    print("#####----------------------------------#####")
    print("--- Execution time: %s seconds ---" % (time.time() - start_time))

